import re
import os
from datetime import datetime

def parse_tasks(notes):
    tasks = []
    for line in notes.strip().split('\n'):
        line = line.strip()
        if not line:
            continue
        task = {
            'date': extract_date(line),
            'description': extract_description(line),
            'hours': extract_hours(line),
            'cost': extract_cost(line),
        }
        if task['description']:
            tasks.append(task)
    return tasks

def extract_date(text):
    m = re.search(r'\b(\d{4}-\d{2}-\d{2})\b', text)
    if m:
        return m.group(1)
    m = re.search(r'\b(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\b', text, re.IGNORECASE)
    if m:
        return m.group(1)
    return datetime.now().strftime('%Y-%m-%d')

def extract_hours(text):
    m = re.search(r'(\d+(?:\.\d+)?)\s*hour', text, re.IGNORECASE)
    return float(m.group(1)) if m else 0.0

def extract_cost(text):
    m = re.search(r'\$\s*(\d+(?:\.\d+)?)', text)
    return float(m.group(1)) if m else 0.0

def extract_description(text):
    text = re.sub(r'\d{4}-\d{2}-\d{2}', '', text)
    text = re.sub(r'\d+(?:\.\d+)?\s*hours?\s*work', '', text, flags=re.IGNORECASE)
    text = re.sub(r'cost\s*\$\d+(?:\.\d+)?', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\$\d+(?:\.\d+)?\s*in\s*materials', '', text, flags=re.IGNORECASE)
    text = re.sub(r'^\s*[>*-]\s*', '', text)
    return text.strip(' ,.')

def generate_reports(notes, customer_name, output_folder):
    tasks = parse_tasks(notes)
    if not tasks:
        raise ValueError("No tasks found. Check your notes format.")
    safe_name = re.sub(r'[^a-zA-Z0-9_]', '_', customer_name).upper()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_name = f"{safe_name}_REPORT_{timestamp}"
    total_hours = sum(t['hours'] for t in tasks)
    total_cost = sum(t['cost'] for t in tasks)
    summary = {
        'tasks': len(tasks),
        'total_hours': total_hours,
        'total_cost': total_cost
    }
    excel_file = generate_excel(tasks, customer_name, summary, output_folder, base_name)
    pdf_file = generate_pdf(tasks, customer_name, summary, output_folder, base_name)
    return {
        'excel': os.path.basename(excel_file),
        'pdf': os.path.basename(pdf_file),
        'summary': summary
    }

def generate_excel(tasks, customer_name, summary, folder, base_name):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Task Report"
        ws.merge_cells('A1:E1')
        ws['A1'] = f"TASK REPORT - {customer_name.upper()}"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill("solid", fgColor="0072FF")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 35
        headers = ['Date', 'Description', 'Hours', 'Cost', 'Notes']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1A1A2E")
            cell.alignment = Alignment(horizontal='center')
        for i, task in enumerate(tasks, 4):
            ws.cell(row=i, column=1, value=task['date'])
            ws.cell(row=i, column=2, value=task['description'])
            ws.cell(row=i, column=3, value=task['hours'])
            ws.cell(row=i, column=4, value=task['cost'])
        total_row = len(tasks) + 4
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=summary['total_hours']).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=summary['total_cost']).font = Font(bold=True)
        for col in range(1, 6):
            ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="FFE066")
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        filepath = os.path.join(folder, f"{base_name}.xlsx")
        wb.save(filepath)
        return filepath
    except ImportError:
        import csv
        filepath = os.path.join(folder, f"{base_name}.csv")
        with open(filepath, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Date', 'Description', 'Hours', 'Cost'])
            for t in tasks:
                writer.writerow([t['date'], t['description'], t['hours'], t['cost']])
            writer.writerow(['TOTAL', '', summary['total_hours'], summary['total_cost']])
        return filepath

def generate_pdf(tasks, customer_name, summary, folder, base_name):
    filepath = os.path.join(folder, f"{base_name}.pdf")
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER
        doc = SimpleDocTemplate(filepath, pagesize=A4,
            topMargin=2*cm, bottomMargin=2*cm,
            leftMargin=2*cm, rightMargin=2*cm)
        elements = []
        blue = colors.HexColor('#0072FF')
        dark = colors.HexColor('#1A1A2E')
        yellow = colors.HexColor('#FFE066')
        title_style = ParagraphStyle('title', fontSize=18,
            textColor=colors.white, backColor=blue,
            alignment=TA_CENTER, spaceAfter=4, leading=28)
        sub_style = ParagraphStyle('sub', fontSize=10,
            textColor=colors.grey, alignment=TA_CENTER, spaceAfter=12)
        elements.append(Paragraph(f"TASK REPORT - {customer_name.upper()}", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y')}", sub_style))
        elements.append(Spacer(1, 0.3*cm))
        table_data = [['Date', 'Description', 'Hours', 'Cost']]
        for t in tasks:
            table_data.append([t['date'], t['description'], str(t['hours']), f"${t['cost']:.2f}"])
        table_data.append(['TOTAL', f"{len(tasks)} tasks", f"{summary['total_hours']} hrs", f"${summary['total_cost']:.2f}"])
        task_table = Table(table_data, colWidths=[3*cm, 9*cm, 2.5*cm, 3*cm], repeatRows=1)
        task_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), dark),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('BACKGROUND', (0,-1), (-1,-1), yellow),
            ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#DDDDDD')),
            ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.HexColor('#E8F4FF'), colors.white]),
        ]))
        elements.append(task_table)
        elements.append(Spacer(1, 1*cm))
        elements.append(Paragraph(
            "Generated by Voice-to-Task | +263 71 141 5728 | +260 774 876 579",
            ParagraphStyle('footer', fontSize=8, textColor=colors.grey, alignment=TA_CENTER)))
        doc.build(elements)
    except ImportError:
        filepath = filepath.replace('.pdf', '.txt')
        with open(filepath, 'w') as f:
            f.write(f"TASK REPORT - {customer_name.upper()}\n{'='*50}\n\n")
            for t in tasks:
                f.write(f"Date: {t['date']}\nTask: {t['description']}\nHours: {t['hours']} | Cost: ${t['cost']}\n\n")
            f.write(f"\nTOTAL: {summary['total_hours']} hrs | ${summary['total_cost']:.2f}\n")
    return filepath
